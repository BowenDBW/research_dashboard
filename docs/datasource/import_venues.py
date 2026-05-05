#!/usr/bin/env python3
"""
初始化刊会分区数据脚本（安全，不删除用户数据）

用途：分发给真实用户时初始化刊会和分区信息
行为：
  - 使用 CREATE TABLE IF NOT EXISTS 创建表（不删除现有表）
  - 使用 INSERT OR IGNORE 插入数据（不覆盖已有数据）
  - 只操作 venues 和 venue_rankings 表，不影响用户数据

数据来源：
  - SCI目录 20260216.csv
  - SSCI目录 20260216.csv
  - 中国计算机学会推荐国际学术会议和期刊目录- 2022.xlsx (CCF)
  - 1750339276-2024JCR完整版.xlsx (JCR)

可重复运行：是，增量导入，不会删除已有数据
"""
import sqlite3
import csv
import io
from pathlib import Path
from openpyxl import load_workbook

def safe_read_csv(path):
    """尝试多种编码读取CSV"""
    encodings = ['utf-8', 'utf-8-sig', 'gbk', 'gb18030', 'latin-1', 'cp1252']
    for enc in encodings:
        try:
            with open(path, 'r', encoding=enc, errors='replace') as f:
                content = f.read()
                return csv.DictReader(io.StringIO(content))
        except Exception:
            continue
    raise Exception(f"无法解码文件: {path}")

def main():
    home_dir = Path.home()
    app_dir = home_dir / ".research_dashboard"
    db_path = app_dir / "research_dashboard.db"
    datasource_dir = Path(__file__).parent

    print(f"[*] 连接数据库: {db_path}")
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # ==========================================
    # 创建刊会表（如果不存在）
    # ==========================================
    print("[*] 创建刊会表...")
    cursor.executescript("""
        -- 刊会基本信息表
        CREATE TABLE IF NOT EXISTS venues (
            venue_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,              -- 刊会名称（可能重复）
            abbreviation TEXT,               -- 简称/缩写
            issn TEXT,                       -- ISSN（期刊唯一标识）
            eissn TEXT,                      -- 电子ISSN
            venue_type TEXT DEFAULT 'journal', -- journal/conference
            publisher TEXT,                  -- 出版社
            url TEXT                         -- 官网/DBLP链接
        );

        -- 刊会分区排名表（一对多：一个刊会可隶属多个分区名单）
        CREATE TABLE IF NOT EXISTS venue_rankings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venue_id INTEGER NOT NULL,
            ranking_source TEXT NOT NULL,    -- ccf/jcr/sci/ssci
            ranking_category TEXT,           -- CCF: A/B/C, JCR: Q1/Q2/Q3/Q4
            ranking_year INTEGER,            -- 分区年份
            category_detail TEXT,            -- 详细分类（如CCF子领域、JCR学科）
            UNIQUE(venue_id, ranking_source, ranking_year)
        );

        -- 创建索引加速查询
        CREATE INDEX IF NOT EXISTS idx_venues_name ON venues(name);
        CREATE INDEX IF NOT EXISTS idx_venues_issn ON venues(issn);
        CREATE INDEX IF NOT EXISTS idx_venues_abbreviation ON venues(abbreviation);
        CREATE INDEX IF NOT EXISTS idx_rankings_venue ON venue_rankings(venue_id);
    """)
    print("[+] 刊会表创建完成")

    # ==========================================
    # 导入 SCI 目录
    # ==========================================
    print("[*] 导入 SCI 目录...")
    sci_path = datasource_dir / "SCI目录 20260216.csv"
    sci_count = 0

    reader = safe_read_csv(sci_path)
    for row in reader:
        name = row.get('Journal title', '').strip()
        issn = row.get('ISSN', '').strip()
        eissn = row.get('eISSN', '').strip()
        publisher = row.get('Publisher name', '').strip()
        categories = row.get('Web of Science Categories', '').strip()

        if not name or not issn:
            continue

        try:
            cursor.execute(
                "INSERT OR IGNORE INTO venues (name, issn, eissn, venue_type, publisher) VALUES (?, ?, ?, 'journal', ?)",
                (name, issn, eissn, publisher)
            )
            if cursor.lastrowid > 0:
                venue_id = cursor.lastrowid
            else:
                cursor.execute("SELECT venue_id FROM venues WHERE issn = ? OR eissn = ?", (issn, eissn or issn))
                result = cursor.fetchone()
                venue_id = result[0] if result else None

            if venue_id:
                cursor.execute(
                    "INSERT OR IGNORE INTO venue_rankings (venue_id, ranking_source, ranking_year, category_detail) VALUES (?, 'sci', 2026, ?)",
                    (venue_id, categories)
                )
                sci_count += 1
        except Exception as e:
            pass

    print(f"[+] SCI 导入完成: {sci_count} 条")

    # ==========================================
    # 导入 SSCI 目录
    # ==========================================
    print("[*] 导入 SSCI 目录...")
    ssci_path = datasource_dir / "SSCI目录 20260216.csv"
    ssci_count = 0

    reader = safe_read_csv(ssci_path)
    for row in reader:
        name = row.get('Journal title', '').strip()
        issn = row.get('ISSN', '').strip()
        eissn = row.get('eISSN', '').strip()
        publisher = row.get('Publisher name', '').strip()
        categories = row.get('Web of Science Categories', '').strip()

        if not name or not issn:
            continue

        try:
            cursor.execute(
                "INSERT OR IGNORE INTO venues (name, issn, eissn, venue_type, publisher) VALUES (?, ?, ?, 'journal', ?)",
                (name, issn, eissn, publisher)
            )
            if cursor.lastrowid > 0:
                venue_id = cursor.lastrowid
            else:
                cursor.execute("SELECT venue_id FROM venues WHERE issn = ? OR eissn = ?", (issn, eissn or issn))
                result = cursor.fetchone()
                venue_id = result[0] if result else None

            if venue_id:
                cursor.execute(
                    "INSERT OR IGNORE INTO venue_rankings (venue_id, ranking_source, ranking_year, category_detail) VALUES (?, 'ssci', 2026, ?)",
                    (venue_id, categories)
                )
                ssci_count += 1
        except Exception as e:
            pass

    print(f"[+] SSCI 导入完成: {ssci_count} 条")

    # ==========================================
    # 导入 CCF 目录
    # ==========================================
    print("[*] 导入 CCF 目录...")
    ccf_path = datasource_dir / "中国计算机学会推荐国际学术会议和期刊目录- 2022.xlsx"
    ccf_count = 0

    wb = load_workbook(ccf_path)
    ws = wb.active

    current_category_group = ""
    current_venue_type = 'journal'  # 默认期刊，根据分组标题动态更新

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue

        # 第一列包含分组标题，判断是会议还是期刊
        if row[0] and isinstance(row[0], str):
            group_title = row[0].replace('\n', ' ').strip()
            current_category_group = group_title
            # 根据标题判断类型
            if '会议' in group_title:
                current_venue_type = 'conference'
            elif '期刊' in group_title:
                current_venue_type = 'journal'

        category = row[1]  # A/B/C
        abbreviation = row[3]  # 刊物简称
        full_name = row[4]  # 刊物全称
        publisher = row[5]  # 出版社
        url = row[6]  # 网址

        if not full_name:
            continue

        # 使用分组标题判断的类型，而不是简称关键词
        venue_type = current_venue_type

        try:
            cursor.execute(
                "INSERT OR IGNORE INTO venues (name, abbreviation, venue_type, publisher, url) VALUES (?, ?, ?, ?, ?)",
                (full_name, abbreviation, venue_type, publisher, url)
            )
            if cursor.lastrowid > 0:
                venue_id = cursor.lastrowid
            else:
                cursor.execute("SELECT venue_id FROM venues WHERE name = ?", (full_name,))
                result = cursor.fetchone()
                venue_id = result[0] if result else None

            if venue_id:
                cursor.execute(
                    "INSERT OR IGNORE INTO venue_rankings (venue_id, ranking_source, ranking_category, ranking_year, category_detail) VALUES (?, 'ccf', ?, 2022, ?)",
                    (venue_id, category, current_category_group)
                )
                ccf_count += 1
        except Exception as e:
            pass

    print(f"[+] CCF 导入完成: {ccf_count} 条")

    # ==========================================
    # 导入 JCR 目录
    # ==========================================
    print("[*] 导入 JCR 目录...")
    jcr_path = datasource_dir / "1750339276-2024JCR完整版.xlsx"
    jcr_count = 0

    wb2 = load_workbook(jcr_path)
    ws2 = wb2.active

    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        name = row[1]
        issn = row[2]
        eissn = row[3]
        category = row[4]
        quartile = row[7]

        if not name or not issn:
            continue

        try:
            cursor.execute(
                "INSERT OR IGNORE INTO venues (name, issn, eissn, venue_type) VALUES (?, ?, ?, 'journal')",
                (name, issn, eissn)
            )
            if cursor.lastrowid > 0:
                venue_id = cursor.lastrowid
            else:
                cursor.execute("SELECT venue_id FROM venues WHERE issn = ? OR eissn = ?", (issn, eissn or issn))
                result = cursor.fetchone()
                venue_id = result[0] if result else None

            if venue_id:
                cursor.execute(
                    "INSERT OR IGNORE INTO venue_rankings (venue_id, ranking_source, ranking_category, ranking_year, category_detail) VALUES (?, 'jcr', ?, 2024, ?)",
                    (venue_id, quartile, category)
                )
                jcr_count += 1
        except Exception as e:
            pass

    print(f"[+] JCR 导入完成: {jcr_count} 条")

    conn.commit()

    # 统计
    cursor.execute("SELECT COUNT(*) FROM venues")
    venue_total = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM venue_rankings")
    ranking_total = cursor.fetchone()[0]

    print("\n[+] ========== 刊会数据导入完成！ ==========")
    print(f"[+] 刊会总数: {venue_total}")
    print(f"[+] 分区记录总数: {ranking_total}")

    conn.close()

if __name__ == "__main__":
    main()